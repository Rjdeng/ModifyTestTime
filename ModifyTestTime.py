import openpyxl
from rich.console import Console
from rich.prompt import Prompt
from datetime import datetime

# 创建一个 Console 实例，用于输出带颜色的文字
console = Console()

# 加载Excel文件
file_path = "appList.xlsx"

def modify_test_time():
    try:
        # 打开 Excel 文件
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook.active

        # 寻找 "测试时间" 列的位置
        test_time_col = None
        for col in sheet.iter_cols(1, sheet.max_column):
            if col[0].value == "测试时间":
                test_time_col = col[0].column
                break

        if test_time_col is None:
            console.print("未找到 '测试时间' 列。", style="bright_yellow")
            return

        # 输入新的测试时间，并确保输入的是0或正整数
        while True:
            console.print("请输入新的测试时间（0 或正整数，分钟）:", style="bold green")
            try:
                new_time = int(Prompt.ask(""))
                if new_time >= 0:
                    break
                else:
                    console.print("输入无效，必须是0或正整数。", style="bold red")
            except ValueError:
                console.print("输入无效，请输入0或正整数。", style="bold red")

        # 更新该列的所有行数据
        for row in range(2, sheet.max_row + 1):  # 从第二行开始，跳过标题行
            sheet.cell(row=row, column=test_time_col, value=new_time)

        # 保存修改后的 Excel 文件
        workbook.save(file_path)

        # 获取当前时间和日期
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        console.print(f"测试时间修改成功！当前时间: {current_time}", style="bold green")

    except Exception as e:
        # 获取当前时间和日期
        current_time = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        console.print(f"修改测试时间失败，当前时间: {current_time}，错误信息: {e}", style="bold red")

    # 提示按任意键退出
    console.print("\n按任意键退出...", style="bold blue")
    input()  # 等待用户按下任意键

if __name__ == "__main__":
    modify_test_time()
